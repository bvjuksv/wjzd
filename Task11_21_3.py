"""
Question for 20221110-17:
Write a program which will read the all the data in the "dataset.xlsx" file.
Then the numbers obtained should be saved sequentially in the first row of a newly generated file named output_order.xlsx.
Tip: Use the python module "openpyxl"
"""

from openpyxl import load_workbook
from openpyxl import Workbook
wb1=Workbook()
ws1=wb1.active
wb2=load_workbook(r'dataset.xlsx')
ws2=wb2.active
data = []

for row in ws2.values:
    row = list(row)
    for item in row:
        data.append(item)
"选择排序"
def select_sort(data):
    m=len(data)-1
    for i in range(m):
        max_index=0
        for j in range(1,len(data)-i):
            if data[j]>data[max_index]:
                max_index=j
            data[m-i],data[max_index]=data[max_index],data[m-i]

    return data

data_order=select_sort(data)
print(data_order)
ws1.append(data_order)
wb1.save(r'output_order3.xlsx')