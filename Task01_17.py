"""读取result_0110中文件提供的total_score、description中的序列（DNA序列）、氨基酸序列并输出到excel表格中"""


from openpyxl import  Workbook
from openpyxl.styles import Font,Alignment

Output_Workbook=Workbook()                                  # 创建一个新的工作簿
sheet1 = Output_Workbook.create_sheet("designed_score",index=0)
sheet2 = Output_Workbook.create_sheet("score_ori_score", index=1)
sheet3 = Output_Workbook.create_sheet("score_refine_score", index=2)
alignment = Alignment(horizontal='center', vertical='center')
font=Font(name='Times New Roman')


with open('designed_score.sc', 'r') as f:                 #读取designed_score文件并存入sheet1
    lines = f.readlines()
    i=1
    for line in lines[1:]:
        columns = line.strip().split()
        if line==lines[1]:
            sheet1.cell(i, 1).value = columns[1]
            sheet1.cell(i, 2).value = columns[18]
        else:
            DNA_list=columns[18].strip().split()
            DNA_sequence="".join(DNA_list)
            i+=1
            sheet1.cell(i,1).value=columns[1]
            sheet1.cell(i,2).value=DNA_sequence[22:32]
with open('designed_MotA3end.fasta', 'r') as f:             #读取多序列比对fasta文件中的氨基酸序列
    lines = f.readlines()
    i=2
    for line in lines[1::3]:
        line_str = "".join(line)
        sheet1.cell(1,3).value="氨基酸序列"
        sheet1.cell(i,3).value=line_str
        i+=1
#设置单元格格式
for row in sheet1.iter_rows(min_row=1, max_row=i+2, min_col=1, max_col=3):      #设置单元格格式
        for cell in row:
            cell.font = font
            cell.alignment = alignment
            sheet1.column_dimensions['A'].width = 15
            sheet1.column_dimensions['B'].width = 20

with open('score_ori_score.sc', 'r') as f:        #读取score_ori_score文件并存入sheet2
    lines = f.readlines()
    i=1
    for line in lines[1:]:
        columns = line.strip().split()
        if line==lines[1]:
            sheet2.cell(i, 1).value = columns[1]
            sheet2.cell(i, 2).value = columns[18]
        else:
            DNA_list=columns[18].strip().split()
            DNA_sequence="".join(DNA_list)
            i+=1
            sheet2.cell(i,1).value=columns[1]
            sheet2.cell(i,2).value=DNA_sequence[10:20]
for row in sheet2.iter_rows(min_row=1, max_row=i, min_col=1, max_col=2):
        for cell in row:
            cell.font = font
            cell.alignment = alignment
            sheet2.column_dimensions['A'].width = 15
            sheet2.column_dimensions['B'].width = 20


with open('score_refine_score.sc', 'r') as f:           #读取score_refine_score文件并存入sheet3
    lines = f.readlines()
    i=1
    for line in lines[1:]:
        columns = line.strip().split()
        if line==lines[1]:
            sheet3.cell(i, 1).value = columns[1]
            sheet3.cell(i, 2).value = columns[18]
        else:
            DNA_list=columns[18].strip().split()
            DNA_sequence="".join(DNA_list)
            i+=1
            sheet3.cell(i,1).value=columns[1]
            sheet3.cell(i,2).value=DNA_sequence[26:36]
for row in sheet3.iter_rows(min_row=1, max_row=i, min_col=1, max_col=2):
        for cell in row:
            cell.font=font
            cell.alignment = alignment
            sheet3.column_dimensions['A'].width = 15
            sheet3.column_dimensions['B'].width = 20

#好像进行了一些没什么用的重复工作反正
print(i)
Output_Workbook.save("collcet.xlsx")
