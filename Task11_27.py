"""Fluorescent Density Calculator
Aim:
Automatically read the raw data of OD and Fluorescence from Excel.
Calculate the fluorescent density and OD, then write it to a new excel.
tips:
Use relative path for input excel.
Program should work fine no matter how many wells are loaded with samples.
Take recent experiment data as a demo input."""

from openpyxl import load_workbook
from openpyxl import Workbook

wb1=load_workbook(r'experiment data.xlsx')
ws1=wb1.active
wb2=Workbook()
ws2=wb2.active
#data=[]

j=1
for row in ws1.iter_rows(min_row=29,max_row=32,min_col=2,max_col=14):
    i=1
    for item in row:
        if bool(item.value)==False:
            i+=1
            continue
        else:
            item=float(item.value)
            item=((item-0.045)*2.6841-0.0006)*8
            item = round(item, 4)
            i += 1
            ws2.cell(1+j,i,item)

    j+=1
print(j)

ws2.cell(1,1).value="OD600"
ws2.cell(j+1,1).value="Fluorescent"

for row in ws1.iter_rows(min_row=53+j,max_row=56+j,min_col=2,max_col=14):
    i=1
    for item in row:
        if bool(item.value)==False:
            i+=1
            continue
        else:
            item=float(item.value)
            item=(item)*8
            item = round(item, 4)
            i += 1
            ws2.cell(2+j,i,item)

    j+=1

wb2.save(r'output.xlsx')