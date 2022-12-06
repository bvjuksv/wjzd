"""
Question for 20221110-17:
Write a program which will read the all the data in the "dataset.xlsx" file.
Then the numbers obtained should be saved sequentially in the first row of a newly generated file named output_order.xlsx.
Tip: Use the python module "openpyxl"
"""

from openpyxl import load_workbook
from openpyxl import Workbook

wb1=Workbook()
ws1=wb1.active

wb2=load_workbook(r'dataset.xlsx')
ws2=wb2.active
data = []

for row in ws2.values:
    row = list(row)
    for item in row:
        data.append(item)
"冒泡排序"
def bubble_sort(data):
    for i in range(len(data)-1):
        for j in range(len(data)-1-i):
            if data[j]>data[j+1]:
                data[j],data[j+1]=data[j+1],data[j]
    return data

data_order=bubble_sort(data)
print(data_order)
ws1.append(data_order)
wb1.save(r'output_order2.xlsx')
