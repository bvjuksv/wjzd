"""
Question for 20221110-17:
Write a program which will read the all the data in the "dataset.xlsx" file.
Then the numbers obtained should be saved in the first row of a newly generated file named output.xlsx.
Tip: Use the python module "openpyxl"
"""

from openpyxl import Workbook
wb1=Workbook()
ws1=wb1.active
i=1
from openpyxl import load_workbook
wb2=load_workbook(r'dataset.xlsx')
ws2=wb2.active
for cell in ws2.values:
    for item in cell:
        print(item)
        ws1.cell(1,i).value=item
        i+=1


wb1.save(r'output1.xlsx')


