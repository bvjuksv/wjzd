"""
Question for 20221110-17:
Write a program which will read the all the data in the "dataset.xlsx" file.
Then the numbers obtained should be saved in the first row of a newly generated file named output.xlsx.
Tip: Use the python module "openpyxl"
"""

from openpyxl import Workbook
wb1=Workbook()
ws1=wb1.active
z=1
from openpyxl import load_workbook
wb2=load_workbook(r'dataset.xlsx')
ws2=wb2.active
for i in range(1,ws2.max_row+1):
    for j in range(1,ws2.max_column+1):
        cell2=ws2.cell(i,j).value
        ws1.cell(1,z).value=cell2
        z+=1

wb1.save(r'output2.xlsx')